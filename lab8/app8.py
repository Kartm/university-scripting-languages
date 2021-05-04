import csv
import argparse
import sys
import openpyxl
from typing import List

# https://www.kaggle.com/rdoume/beerreviews

parser = argparse.ArgumentParser(
    description="Program for analyzing beer reviews",
    add_help=True
)
parser.add_argument(
    "path",
    help="dataset path",
)
parser.add_argument("-o", "--output",
                    help="name of excel file to be generated",
                    )


def get_dataset_path(args):
    path = args.path
    if len(path) > 1:
        if path.endswith(".csv"):
            return path
        print("Dataset path must end with \".csv\".")
        sys.exit(0)
    print("Dataset path must be specified.")
    sys.exit(0)


class BeerReview:
    def __init__(self, brewery_name, review_overall, review_profilename, beer_style):
        self.brewery_name = brewery_name
        self.review_overall = review_overall
        self.review_profilename = review_profilename
        self.beer_style = beer_style

    def __str__(self):
        return f"{self.brewery_name}'s {self.beer_style} was rated {self.review_overall} by {self.review_profilename}"

    def __repr__(self):
        return f"{self.brewery_name}'s {self.beer_style} was rated {self.review_overall} by {self.review_profilename}"


def read_csv_file(path: str):
    i = 0

    rows = []
    with open(path) as f:
        reader = csv.reader(f)
        next(reader)  # skip header

        for row in reader:
            rows.append(BeerReview(brewery_name=row[1],
                                   review_overall=float(row[3]),
                                   review_profilename=row[6],
                                   beer_style=row[7]
                                   ))

            i += 1
            if i == 100:
                break

    return rows


def read_dataset(path: str):
    try:
        return read_csv_file(path)
    except FileNotFoundError:
        print('Error: Log file not found.')
        sys.exit(0)


def get_stats(reviews: List[BeerReview]):
    stats = dict()

    mean_rating = sum(review.review_overall for review in reviews) / len(reviews)
    stats['mean_rating'] = mean_rating

    worst_ratings_of_users = dict()
    for review in reviews:
        if worst_ratings_of_users.get(review.review_profilename):
            worst_ratings_of_users[review.review_profilename] = min(
                review.review_overall,
                worst_ratings_of_users.get(review.review_profilename)
            )
        else:
            worst_ratings_of_users[review.review_profilename] = review.review_overall

    stats['worst_ratings_of_users'] = worst_ratings_of_users
    stats['all_reviews_count'] = len(reviews)

    return stats


def save_to_excel(excel_path, stats):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Beer reviews analysis"

    f = openpyxl.styles.Font(color="FFD700", italic=True, bold=True)

    sheet.cell(row=2, column=2, value="Mean rating")
    sheet.cell(row=3, column=2, value="Worst ratings of users")
    sheet.cell(row=4, column=2, value="How many reviews in total")

    sheet.cell(row=2, column=3, value=stats.get('mean_rating')).font = f
    sheet.cell(row=3, column=3, value=str(stats.get('worst_ratings_of_users'))).font = f
    sheet.cell(row=4, column=3, value=stats.get('all_reviews_count')).font = f

    workbook.save(excel_path)


def run():
    args = parser.parse_args()

    dataset_path = get_dataset_path(args)
    excel_path = args.output

    reviews = read_dataset(dataset_path)
    # print_stats(dataset)
    # print(dataset)

    stats = get_stats(reviews)
    if excel_path:
        save_to_excel(excel_path, stats)
    else:
        print(f"Mean rating: {stats.get('mean_rating')}")
        print(f"Worst ratings of users: {stats.get('worst_ratings_of_users')}")
        print(f"How many reviews in total: {stats.get('all_reviews_count')}")
        # print(stats)


if __name__ == "__main__":
    run()
